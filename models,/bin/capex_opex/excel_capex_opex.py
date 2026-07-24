from __future__ import annotations
from pathlib import Path
from typing import Dict, Tuple, Optional
from openpyxl import load_workbook
import pandas as pd

def calc_capex_opex(
    excel_path: str | Path,
    inputs: Dict[str, float],
    return_breakdowns: bool = False,
    ref_cost_overrides: Optional[Dict[str, float]] = None,  # <- NEW (keys in MUSD)
) -> Tuple[float, float, Optional[pd.DataFrame], Optional[pd.DataFrame]]:
    """
    Compute Total CAPEX (MUSD) and Total OPEX (MUSD/year) from the Excel model.

    New:
      ref_cost_overrides (optional): dict to override CAPEX reference costs (Dref, MUSD).
        Supported keys (case-insensitive aliases in parentheses):
          - "BOP"                 ( "CSP BOP" )
          - "SF_aperture"         ( "Solar Field Aperture Area", "CSP SF" )
          - "PV_modules"          ( "PV Modules" )

        Example:
          ref_cost_overrides={
              "BOP": 325.0,          # MUSD
              "SF_aperture": 480.0,  # MUSD
              "PV_modules": 210.0    # MUSD
          }
    """
    excel_path = Path(excel_path)

    # Load twice: formulas (structure) and values (numbers)
    wbF = load_workbook(excel_path, data_only=False)
    wbN = load_workbook(excel_path, data_only=True)
    capF, capN = wbF["CAPEX"], wbN["CAPEX"]
    oxF,  oxN  = wbF["OPEX"],  wbN["OPEX"]

    # -------- CAPEX references (rows 26..45) --------
    capex_refs = {}
    for r in range(26, 46):
        lbl = capF.cell(r,1).value
        Cref = capF.cell(r,3).value
        Dref = capN.cell(r,4).value  # numeric (MUSD)
        if Cref is None or Dref is None:
            raise ValueError(f"Missing CAPEX reference at row {r}: C={Cref}, D={Dref}")
        capex_refs[r] = (lbl, float(Cref), float(Dref))

    # --- Apply optional reference-cost overrides (in MUSD) ---
    if ref_cost_overrides:
        # Normalize keys
        keymap = {
            "bop": 27,                   # CSP BOP
            "csp bop": 27,
            "sf_aperture": 28,           # Solar Field Aperture Area (Mirror Area)
            "solar field aperture area": 28,
            "csp sf": 28,
            "pv_modules": 41,            # PV Modules
            "pv modules": 41,
        }
        for k_raw, v in ref_cost_overrides.items():
            k = str(k_raw).strip().lower()
            if k in keymap:
                row = keymap[k]
            else:
                # try fuzzy: strip spaces/underscores
                k2 = k.replace(" ", "_")
                row = keymap.get(k2)
            if row is None:
                continue  # ignore unsupported keys silently
            lbl, Cref, _old_Dref = capex_refs[row]
            capex_refs[row] = (lbl, Cref, float(v))  # override Dref (MUSD)

    def capex_power_cost(ref_row: int, size: float, coeff_cell_val) -> float:
        # Excel algebra: D = D_ref * (size / C_ref) ^ coeff
        _, Cref, Dref = capex_refs[ref_row]
        coeff = float(coeff_cell_val)
        return Dref * (float(size) / Cref)**coeff  # MUSD

    # Helper: sizes
    get = inputs.__getitem__  # convenience alias

    # CAPEX rows 50..69 (build component costs using references)
    rows = []
    rows.append(("CSP Power Block", capex_power_cost(26, get("PB Installed Capacity (Gross)"), capF.cell(50,2).value)))
    rows.append(("CSP BOP",         capex_power_cost(27, get("PB Installed Capacity (Gross)"), capF.cell(51,2).value)))
    rows.append(("CSP TES",         capex_power_cost(29, get("Thermal Energy Storage Capacity "), capF.cell(52,2).value)))
    rows.append(("CSP SF",          capex_power_cost(28, get("Solar Field Aperture Area (Mirror Area)"), capF.cell(53,2).value)))
    rows.append(("CSP Receiver",    capex_power_cost(30, get("Receiver Power (Max Rated)"), capF.cell(54,2).value)))
    rows.append(("CSP Tower",       capex_power_cost(31, get("Tower Height (w/o Receiver)"), capF.cell(55,2).value)))
    rows.append(("CSP Site/Civil",  capex_power_cost(32, get("Land Area CSP"), capF.cell(56,2).value)))

    # Electric Heater (linear): = D33/C33 * size
    _, C33, D33 = capex_refs[33]
    rows.append(("Electric Heater", (D33 / C33) * get("Electric Heater Thermal Power (Max Rated)")))

    rows.append(("Grid Interconnection", capex_power_cost(34, get("Distance to Grid "), capF.cell(58,2).value)))

    # Road (Excel sets C59=2 km fixed)
    rows.append(("Road Costs", capex_power_cost(35, 2, capF.cell(59,2).value)))

    rows.append(("Gas Pipeline",    capex_power_cost(36, get("Distance to Gas"), capF.cell(60,2).value)))
    rows.append(("Water Pipeline",  capex_power_cost(37, get("Distance to Water"), capF.cell(61,2).value)))
    rows.append(("Battery Power",   capex_power_cost(38, get("Battery Pack Power (Max Rated)"), capF.cell(62,2).value)))
    rows.append(("Battery Energy",  capex_power_cost(39, get("Battery Pack Capacity"), capF.cell(63,2).value)))
    rows.append(("Other Component", capex_power_cost(40, get("Other Component Size"), capF.cell(64,2).value)))

    # PV sub-blocks (PV Installed Capacity in Wdc)
    pv_wdc = get("PV Installed Capacity")
    rows.append(("PV Modules",     capex_power_cost(41, pv_wdc, capF.cell(65,2).value)))
    rows.append(("PV Inverters",   capex_power_cost(42, pv_wdc, capF.cell(66,2).value)))
    rows.append(("PV BOS",         capex_power_cost(43, pv_wdc, capF.cell(67,2).value)))
    rows.append(("PV SA Tracking", capex_power_cost(44, pv_wdc, capF.cell(68,2).value)))
    rows.append(("PV Site/Civil",  capex_power_cost(45, get("Land Area PV"), capF.cell(69,2).value)))

    capex_df = pd.DataFrame(rows, columns=["Component", "MUSD"])

    # Totals (rows 71..78)
    capex_direct_subtotal = capex_df["MUSD"].sum()
    contingency_rate = float(capF.cell(71,3).value)  # C71
    direct_total = capex_direct_subtotal * (1 + contingency_rate)

    indirect_rate = float(capF.cell(74,3).value) + float(capF.cell(75,3).value) + float(capF.cell(76,3).value)
    indirect_total = direct_total * indirect_rate
    total_capex_musd = direct_total + indirect_total

    # -------- OPEX references (rows 27..44) --------
    opex_refs = {}
    for r in range(27,45):
        lbl = oxF.cell(r,1).value
        Cref = oxN.cell(r,3).value  # numeric
        Dref = oxN.cell(r,4).value  # numeric USD
        if isinstance(Cref, (int,float)) and isinstance(Dref, (int,float)):
            opex_refs[r] = (lbl, float(Cref), float(Dref))

    def opex_power_cost(ref_row: int, size: float, coeff_cell_val) -> float:
        # Excel algebra: USD = D_ref * (size / C_ref) ^ coeff
        if ref_row not in opex_refs:
            raise ValueError(f"OPEX reference row {ref_row} missing numeric values")
        _, Cref, Dref = opex_refs[ref_row]
        coeff = float(coeff_cell_val)
        return Dref * (float(size) / Cref)**coeff

    sizes = {
        "C4":  inputs["PB Installed Capacity (Gross)"],
        "C5":  inputs["Solar Field Aperture Area (Mirror Area)"],
        "C12": inputs["Battery Pack Power (Max Rated)"],
        "C14": inputs["Battery Annual Generation (for OPEX)"],
        "C16": inputs["CSP Annual Generation (for OPEX)"],
        "C21": inputs["Other Component Size"],
        "C11": inputs["PV Installed Capacity"],
        "C15": inputs["Land Area PV"],
    }

    ox_items = []
    # CSP
    ox_items.append(("CSP Admin/Finance & Mgmt",   opex_power_cost(27, sizes["C4"],  oxF.cell(50,2).value)))
    ox_items.append(("CSP Plant Ops Labor",        opex_power_cost(28, sizes["C4"],  oxF.cell(51,2).value)))
    ox_items.append(("CSP Plant Maint Labor",      opex_power_cost(29, sizes["C4"],  oxF.cell(52,2).value)))
    ox_items.append(("CSP SF Maint Labor",         opex_power_cost(30, sizes["C5"],  oxF.cell(53,2).value)))
    ox_items.append(("CSP Subcontracted Services", opex_power_cost(33, sizes["C16"], oxF.cell(54,2).value)))
    ox_items.append(("CSP Electricity Use",        opex_power_cost(38, sizes["C4"],  oxF.cell(55,2).value)))
    ox_items.append(("CSP Water Use",              opex_power_cost(39, sizes["C4"],  oxF.cell(56,2).value)))
    ox_items.append(("CSP Machinery",              opex_power_cost(43, sizes["C4"],  oxF.cell(57,2).value)))
    ox_items.append(("CSP Spare Parts",            opex_power_cost(44, sizes["C16"], oxF.cell(58,2).value)))
    # BESS
    ox_items.append(("BESS Fixed",                 opex_power_cost(34, sizes["C12"], oxF.cell(59,2).value)))
    ox_items.append(("BESS Variable",              opex_power_cost(35, sizes["C14"], oxF.cell(60,2).value)))
    # Other
    ox_items.append(("Other OPEX",                 opex_power_cost(42, sizes["C21"], oxF.cell(61,2).value)))
    # PV fixed OPEX (row 62): SUM(CAPEX!D65:D69)*1e6*G6/100
    pv_om_percent = float(ref_cost_overrides.get("pv_fixed_opex_coeff", oxN.cell(6, 7).value))
    pv_capex_usd  = sum(float(capN.cell(r,4).value) for r in range(65,70)) * 1_000_000.0
    ox_items.append(("PV Fixed OPEX", pv_capex_usd * pv_om_percent / 100.0))

    opex_df = pd.DataFrame(ox_items, columns=["Item","USD"]).sort_values("USD", ascending=False)
    total_opex_musd = opex_df["USD"].sum() / 1e6

    if return_breakdowns:
        return total_capex_musd, total_opex_musd, capex_df, opex_df
    else:
        return total_capex_musd, total_opex_musd, None, None
