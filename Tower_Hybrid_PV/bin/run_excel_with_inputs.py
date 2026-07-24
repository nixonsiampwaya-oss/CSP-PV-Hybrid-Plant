from pathlib import Path
from openpyxl import load_workbook

# ---- File path ----
HERE = Path(__file__).parent
EXCEL_PATH = HERE / "CAPEX OPEX model CSP-PV MJ2500.xlsx"

# ---- Your inputs (edit as needed) ----
inputs = {
    "PB Installed Capacity (Gross)": 200000,
    "Solar Field Aperture Area (Mirror Area)": 507597.5,
    "Thermal Energy Storage Capacity ": 3271.3,
    "Receiver Power (Max Rated)": 200,
    "Tower Height (w/o Receiver)": 200,
    "Electric Heater Thermal Power (Max Rated)": 200000,
    "Land Area CSP": 2_000_000,
    "PV Installed Capacity": 270_327_760,
    "Battery Pack Power (Max Rated)": 150,
    "Battery Pack Capacity": 300,
    "Battery Annual Generation (for OPEX)": 400,
    "Land Area PV": 1_000_000,
    "CSP Annual Generation (for OPEX)": 213.85,
    "Distance to Grid ": 4,
    "Distance to Road": 3,
    "Distance to Gas": 5,
    "Distance to Water": 10,
    "Other Component Size": 0,
}

# ---- Write inputs to Excel ----
def write_inputs(file_path: Path, inputs: dict) -> dict:
    wb = load_workbook(file_path, data_only=False)
    ws = wb["INPUTS - SUMMARY"]
    written = {}
    for r in range(1, ws.max_row + 1):
        label = ws.cell(r, 1).value  # Column A
        if isinstance(label, str):
            for key, val in inputs.items():
                if label.strip().startswith(key):
                    ws.cell(r, 3, value=val)  # Column C holds the numeric value
                    written[key] = (r, val)
    wb.save(file_path)
    return written

# ---- Recalculate Excel formulas via xlwings (if available) ----
def recalc_with_excel(file_path: Path) -> bool:
    try:
        import xlwings as xw
        app = xw.App(visible=False, add_book=False)
        book = xw.Book(file_path.as_posix())
        book.app.api.CalculateFull()
        book.save()
        book.close()
        app.quit()
        return True
    except Exception as e:
        print(f"[!] Could not recalc with Excel via xlwings: {e}")
        print("    -> Open the workbook manually and press Ctrl+Alt+F9 to force full recalculation, then save.")
        return False

# ---- Read CAPEX and OPEX results ----
def read_totals(file_path: Path) -> tuple[float, float]:
    wb = load_workbook(file_path, data_only=True)
    ws_in = wb["INPUTS - SUMMARY"]
    ws_ox = wb["OPEX"]

    capex_musd = None
    for r in range(1, ws_in.max_row + 1):
        a = ws_in.cell(r, 1).value
        if isinstance(a, str) and a.strip().upper().startswith("TOTAL CAPEX"):
            capex_musd = ws_in.cell(r, 2).value  # Column B is MUSD
            break

    opex_musd = None
    for r in range(1, ws_ox.max_row + 1):
        a = ws_ox.cell(r, 1).value
        if isinstance(a, str) and "OPEX TOTAL" in a.upper():
            opex_musd = ws_ox.cell(r, 4).value  # Column D is MUSD
            break

    return capex_musd, opex_musd

# ---- Main function ----
def run_excel_model():
    if not EXCEL_PATH.exists():
        raise FileNotFoundError(f"Excel model not found: {EXCEL_PATH}")

    print(f"\nUpdating inputs directly in: {EXCEL_PATH.name}")
    written = write_inputs(EXCEL_PATH, inputs)
    did_calc = recalc_with_excel(EXCEL_PATH)
    capex_musd, opex_musd = read_totals(EXCEL_PATH)

    print("\n=== Hybrid CSP–PV CAPEX/OPEX (from Excel model) ===")
    print(f"Recalculated with Excel via xlwings: {'YES' if did_calc else 'NO'}")
    if capex_musd is not None:
        print(f"Total CAPEX : {capex_musd:,.3f} MUSD")
        print(f"              ≈ {capex_musd*1e6:,.0f} USD")
    else:
        print("Total CAPEX : (not found — open Excel and press Ctrl+Alt+F9)")

    if opex_musd is not None:
        print(f"Total OPEX  : {opex_musd:,.3f} MUSD/year")
        print(f"              ≈ {opex_musd*1e6:,.0f} USD/year")
    else:
        print("Total OPEX  : (not found — open Excel and press Ctrl+Alt+F9)")

    return capex_musd, opex_musd

# ---- Run directly (for Jupyter) ----
capex, opex = run_excel_model()

print(f"\nReturned values:\nCAPEX = {capex} MUSD\nOPEX = {opex} MUSD/year")
